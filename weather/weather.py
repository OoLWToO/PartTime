import re

import requests
from lxml import etree
from matplotlib import pyplot as plt

# 设置请求头，用于模拟用户真实行为
from openpyxl import Workbook

headers = {
    'accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9',
    'accept-encoding': 'gzip, deflate, br',
    'accept-language': 'zh-CN,zh;q=0.9',
    'cookie': 'Hm_lvt_e9daa3d2d6076d2d8ec79d05f050c0d5=1716811706,1716825458; Hm_lvt_029c70397ba7bba8caeb29017c83b8d8=1716811706,1716825458; Hm_lpvt_029c70397ba7bba8caeb29017c83b8d8=1716825462; Hm_lpvt_e9daa3d2d6076d2d8ec79d05f050c0d5=1716825462',
    'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/107.0.0.0 Safari/537.36',
}


# 创建折线图
def create_broken_line_chart(x, y, title):
    y = [int(re.search(r'~(\d+)℃', temp).group(1)) for temp in y]
    plt.figure(figsize=(20, 7))
    plt.plot(x, y, marker='o')
    plt.xticks(x, x, rotation=90, fontsize=8)
    plt.rcParams['font.sans-serif'] = ['SimHei']
    plt.title(title)
    plt.savefig(f'{title}折线图.png')


# 创建条形图
def create_line_chart(y, title):
    item = []
    value = []
    for i in y:
        if i not in item:
            item.append(i)
            value.append(1)
        else:
            value[item.index(i)] += 1
    sorted_item = [x for _, x in sorted(zip(value, item), reverse=True)]
    sorted_value = sorted(value, reverse=True)
    plt.bar(sorted_item, sorted_value)
    plt.rcParams['font.sans-serif'] = ['SimHei']
    plt.title(title)
    plt.savefig(f'{title}条形图.png')


# 创建饼图
def create_pie_chart(situation, title):
    item = []
    value = []
    # 统计数据
    for i in situation:
        if i not in item:
            item.append(i)
            value.append(1)
        else:
            value[item.index(i)] += 1
    sorted_item = [x for _, x in sorted(zip(value, item), reverse=True)]
    sorted_value = sorted(value, reverse=True)
    plt.pie(sorted_value, labels=sorted_item, autopct='%1.1f%%', startangle=140)
    plt.rcParams['font.sans-serif'] = ['SimHei']
    plt.axis('equal')
    plt.title(title)
    plt.savefig(f'{title}饼图.png')


if __name__ == '__main__':
    # 设置爬取的城市
    city_urls = ['shanghai', 'beijing', 'tianjin']
    city_name = ['上海', '北京', '天津']
    for i in range(len(city_urls)):
        # 请求url
        url = f'https://www.tianqishi.com/lishi/{city_urls[i]}.html'
        r = requests.get(url, headers=headers)
        html = etree.HTML(r.text)
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.column_dimensions['A'].width = 15
        worksheet.cell(row=1, column=1, value='时间')
        worksheet.cell(row=1, column=2, value='气温')
        worksheet.cell(row=1, column=3, value='天气情况')
        worksheet.cell(row=1, column=4, value='风向')
        worksheet.cell(row=1, column=5, value='风力')
        worksheet.cell(row=1, column=6, value='日出')
        worksheet.cell(row=1, column=7, value='日落')
        rows = 2
        time = []
        temperature = []
        situation = []
        wind_direction = []
        wind_power = []
        sunrise = []
        sunset = []
        # 获取天气列表
        weather_list = html.xpath('//*[@class="yuBaoTable"]//tr')
        for weather in weather_list:
            # 获取时间、温度、风向、风速、湿度、下雨概率
            time.append(weather.xpath('./td[1]/a/text()')[0])
            temperature.append(weather.xpath('./td[2]/text()')[0])
            situation.append(weather.xpath('./td[3]/text()')[0])
            wind_direction.append(weather.xpath('./td[4]/text()')[0])
            wind_power.append(weather.xpath('./td[5]/text()')[0])
            sunrise.append(weather.xpath('./td[6]/text()')[0])
            sunset.append(weather.xpath('./td[7]/text()')[0])
            worksheet.cell(row=rows, column=1, value=weather.xpath('./td[1]/a/text()')[0])
            worksheet.cell(row=rows, column=2, value=weather.xpath('./td[2]/text()')[0])
            worksheet.cell(row=rows, column=3, value=weather.xpath('./td[3]/text()')[0])
            worksheet.cell(row=rows, column=4, value=weather.xpath('./td[4]/text()')[0])
            worksheet.cell(row=rows, column=5, value=weather.xpath('./td[5]/text()')[0])
            worksheet.cell(row=rows, column=6, value=weather.xpath('./td[6]/text()')[0])
            worksheet.cell(row=rows, column=7, value=weather.xpath('./td[7]/text()')[0])
            rows += 1
        # 保存到Excel
        workbook.save(f'{city_name[i]}历史天气统计.xlsx')
        # 生成图表
        create_line_chart(wind_direction, f'{city_name[i]}历史风向统计')
        create_broken_line_chart(time, temperature, f'{city_name[i]}历史最高气温统计')
        create_pie_chart(situation, f'{city_name[i]}天气情况统计')
