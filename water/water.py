import requests
from lxml import etree
from openpyxl.workbook import Workbook

headers = {
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7',
    'Accept-Encoding': 'gzip, deflate, br, zstd',
    'accept-language': 'zh-CN,zh;q=0.9',
    'Cache-Control': 'max-age=0',
    'Cookie': '__jsluid_s=c9bf2381199315a3f9d70fab5efbea9b; __jsl_clearance_s=1718590950.375|0|BNVMH65p0N9580f3nELVxfBlj1g%3D; PHPSESSID=mnnnrrlijttt606feek3bkdvg0; mfw_uuid=666f9de7-be92-5b8a-42db-f09d461618a6; oad_n=a%3A3%3A%7Bs%3A3%3A%22oid%22%3Bi%3A1029%3Bs%3A2%3A%22dm%22%3Bs%3A15%3A%22www.mafengwo.cn%22%3Bs%3A2%3A%22ft%22%3Bs%3A19%3A%222024-06-17+10%3A22%3A31%22%3B%7D; __mfwc=direct; __mfwa=1718590951232.55084.1.1718590951232.1718590951232; __mfwlv=1718590951; __mfwvn=1; Hm_lvt_8288b2ed37e5bc9b4c9f7008798d2de0=1718590951; uva=s%3A150%3A%22a%3A3%3A%7Bs%3A2%3A%22lt%22%3Bi%3A1718590952%3Bs%3A10%3A%22last_refer%22%3Bs%3A82%3A%22https%3A%2F%2Fwww.mafengwo.cn%2Flocaldeals%2F0-0-M10132-0-0-0-0-0.html%3Ffrom%3Dlocaldeals_index%22%3Bs%3A5%3A%22rhost%22%3BN%3B%7D%22%3B; __mfwurd=a%3A3%3A%7Bs%3A6%3A%22f_time%22%3Bi%3A1718590952%3Bs%3A9%3A%22f_rdomain%22%3Bs%3A15%3A%22www.mafengwo.cn%22%3Bs%3A6%3A%22f_host%22%3Bs%3A3%3A%22www%22%3B%7D; __mfwuuid=666f9de7-be92-5b8a-42db-f09d461618a6; __mfwb=7d6d46205828.2.direct; __mfwlt=1718590972; Hm_lpvt_8288b2ed37e5bc9b4c9f7008798d2de0=1718590973; bottom_ad_status=0',
    'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/119.0.0.0 Safari/537.36',
}

provinces = ['北京', '天津', '河北', '山西', '内蒙', '辽宁', '吉林', '黑龙江',
             '上海', '江苏', '浙江', '安徽', '福建', '江西', '山东', '河南', '湖北', '湖南', '广东', '广西', '海南',
             '重庆', '四川', '贵州', '云南', '西藏', '陕西', '甘肃', '青海', '宁夏', '新疆', '台湾', '香港', '澳门']

data = {
    '省份': [],
    '城市': [],
    '日期': [],
    '第一阶梯': [],
    '第二阶梯': [],
    '第三阶梯': [],
    '无阶梯': [],
    '非居民': [],
    '特种行业': [],
    '居民（污）': [],
    '非居民（污）': [],
    '特种行业（污）': []
}

if __name__ == '__main__':
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.cell(row=1, column=1, value='省份')
    worksheet.cell(row=1, column=2, value='城市')
    worksheet.cell(row=1, column=3, value='日期')
    worksheet.cell(row=1, column=4, value='第一阶梯')
    worksheet.cell(row=1, column=5, value='第二阶梯')
    worksheet.cell(row=1, column=6, value='第三阶梯')
    worksheet.cell(row=1, column=7, value='无阶梯')
    worksheet.cell(row=1, column=8, value='非居民')
    worksheet.cell(row=1, column=9, value='特种行业')
    worksheet.cell(row=1, column=10, value='居民（污）')
    worksheet.cell(row=1, column=11, value='非居民（污）')
    worksheet.cell(row=1, column=12, value='特种行业（污）')
    rows = 2
    for i in range(0, 34):
        url = f'https://www.h2o-china.com/price/list?pid={i + 1}'
        r = requests.get(url, headers=headers)
        html = etree.HTML(r.text)
        cities = html.xpath('//*[@id="price_c02_l"]//li/a')
        for city in cities:
            for year in range(2015, 2025):
                url = f'https://www.h2o-china.com/price/{city.xpath("./@href")[0]}&ayear={year}'
                print(url)
                r = requests.get(url, headers=headers)
                html = etree.HTML(r.text)
                price_list = html.xpath('//*[@class="price_buss"]/tr')
                for price in price_list:
                    time = price.xpath('./td[1]/strong/text()')[0]
                    water_one = price.xpath('./td[2]/text()')[0].replace('\n', '').replace('\t', '').replace('\r', '')
                    water_two = price.xpath('./td[3]/text()')[0].replace('\n', '').replace('\t', '').replace('\r', '')
                    water_three = price.xpath('./td[4]/text()')[0].replace('\n', '').replace('\t', '').replace('\r', '')
                    water_not = price.xpath('./td[5]/text()')[0].replace('\n', '').replace('\t', '').replace('\r', '')
                    water_no_jm = price.xpath('./td[6]/text()')[0].replace('\n', '').replace('\t', '').replace('\r', '')
                    water_tzhy = price.xpath('./td[7]/text()')[0].replace('\n', '').replace('\t', '').replace('\r', '')
                    dirty_water_jm = price.xpath('./td[8]/text()')[0].replace('\n', '').replace('\t', '').replace('\r', '')
                    dirty_water_no_jm = price.xpath('./td[9]/text()')[0].replace('\n', '').replace('\t', '').replace('\r', '')
                    dirty_water_tzhy = price.xpath('./td[10]/text()')[0].replace('\n', '').replace('\t', '').replace('\r', '')
                    if water_one != '' or water_two != '' or water_three != '' or water_not != '' or water_no_jm != '' or water_tzhy != '' or dirty_water_jm != '' or dirty_water_no_jm != '' or dirty_water_tzhy != '':
                        print(f'{provinces[i]}   {city.text}   {time}   {water_one}   {water_two}   {water_three}   {water_not}   {water_no_jm}   {water_tzhy}   {dirty_water_jm}   {dirty_water_no_jm}   {dirty_water_tzhy}')
                        worksheet.cell(row=rows, column=1, value=provinces[i])
                        worksheet.cell(row=rows, column=2, value=city.text)
                        worksheet.cell(row=rows, column=3, value=time)
                        worksheet.cell(row=rows, column=4, value=water_one)
                        worksheet.cell(row=rows, column=5, value=water_two)
                        worksheet.cell(row=rows, column=6, value=water_three)
                        worksheet.cell(row=rows, column=7, value=water_not)
                        worksheet.cell(row=rows, column=8, value=water_no_jm)
                        worksheet.cell(row=rows, column=9, value=water_tzhy)
                        worksheet.cell(row=rows, column=10, value=dirty_water_jm)
                        worksheet.cell(row=rows, column=11, value=dirty_water_no_jm)
                        worksheet.cell(row=rows, column=12, value=dirty_water_tzhy)
                        rows += 1
        workbook.save('中国水网水价数据.xlsx')

