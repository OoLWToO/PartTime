import pandas as pd
from openpyxl import Workbook
from pyecharts.faker import Faker
from pyecharts.globals import SymbolType

from selenium import webdriver
from pyecharts import options as opts
from pyecharts.charts import Line, Bar, Pie, WordCloud


def create_driver():
    chrome_driver_path = '../driver/chromedriver_119.exe'
    option = webdriver.ChromeOptions()
    option.add_experimental_option("detach", True)
    driver = webdriver.Chrome(executable_path=chrome_driver_path, chrome_options=option)
    driver.implicitly_wait(.5)
    driver.maximize_window()
    return driver


def data_statistics(pd_data, type):
    company_item = []
    company_value = []
    for data in pd_data[type]:
        if data not in company_item:
            company_item.append(data)
            company_value.append(1)
        else:
            company_value[company_item.index(data)] += 1
    sorted_data = [x for _, x in sorted(zip(company_value, company_item), reverse=True)]
    sorted_value = sorted(company_value, reverse=True)
    return [sorted_data, sorted_value]


# 公司类型柱状图
def create_line_chart(pd_data):
    statistics = data_statistics(pd_data, '公司类型')
    company_type_item = statistics[0]
    company_type_value = statistics[1]
    index_to_remove = company_type_item.index('-')
    del company_type_item[index_to_remove]
    del company_type_value[index_to_remove]
    bar = Bar()
    bar.add_xaxis(company_type_item)
    bar.add_yaxis("数量", company_type_value)
    bar.render("company_type.html")


# 工作地点饼图
def create_pie_chart(pd_data):
    statistics = data_statistics(pd_data, '工作地点')
    company_address_item = statistics[0]
    company_address_value = statistics[1]
    pie = (
        Pie()
        .add("", [list(z) for z in zip(company_address_item, company_address_value)])
        .set_global_opts(title_opts=opts.TitleOpts(title="工作地点统计"))
        .set_series_opts(label_opts=opts.LabelOpts(formatter="{b}:{c} {d}%"))
    )
    pie.render("company_address.html")


# 经营项目词云
def create_word_chart(pd_data):
    statistics = data_statistics(pd_data, '经营项目')
    company_business_item = statistics[0]
    company_business_value = statistics[1]
    index_to_remove = company_business_item.index('-')
    del company_business_item[index_to_remove]
    del company_business_value[index_to_remove]
    business_data = list(zip(company_business_item, company_business_value))
    word = (
        WordCloud()
        .add("经营项目", business_data, word_size_range=[20, 100], shape=SymbolType.DIAMOND)
        .set_global_opts(title_opts=opts.TitleOpts(title="公司经营项目统计"))
    )
    word.render("company_business.html")


if __name__ == "__main__":
    driver = create_driver()
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.column_dimensions['A'].width = 30
    worksheet.column_dimensions['B'].width = 14
    worksheet.column_dimensions['C'].width = 18
    worksheet.column_dimensions['D'].width = 40
    worksheet.column_dimensions['E'].width = 12
    worksheet.column_dimensions['F'].width = 12
    worksheet.column_dimensions['G'].width = 30
    worksheet.column_dimensions['H'].width = 10
    worksheet.cell(row=1, column=1, value='职位名称')
    worksheet.cell(row=1, column=2, value='职位月薪')
    worksheet.cell(row=1, column=3, value='工作地点')
    worksheet.cell(row=1, column=4, value='公司名称')
    worksheet.cell(row=1, column=5, value='公司类型')
    worksheet.cell(row=1, column=6, value='公司规模')
    worksheet.cell(row=1, column=7, value='经营项目')
    worksheet.cell(row=1, column=8, value='其他信息')
    worksheet.cell(row=1, column=9, value='刷新日期')
    data = {
        '职位名称': [],
        '职位月薪': [],
        '工作地点': [],
        '公司名称': [],
        '公司类型': [],
        '公司规模': [],
        '经营项目': [],
        '其他信息': [],
        '刷新日期': []
    }
    rows = 2
    for page in range(1, 30):
        driver.get(f'https://www.job001.cn/jobs?keyType=0&keyWord=&jobTypeId=&jobType=%E8%81%8C%E4%BD%8D%E7%B1%BB%E5%9E%8B&industry=&industryname=%E8%A1%8C%E4%B8%9A%E7%B1%BB%E5%9E%8B&workId=25.293&workPlace=%E7%8F%A0%E6%B5%B7%E5%B8%82&salary=&salaryType=&entType=&experience=&education=&entSize=&benefits=&reftime=&workTypeId=&sortField=&pageNo={page}&curItem=&searchType=1')
        jobs_list = driver.find_elements_by_xpath('//*[@class="jobsList"]')
        for job in jobs_list:
            job_name = job.find_element_by_xpath('.//*[@class="jobNameCon"]').text.replace(' ', '').replace('\n', '')
            salary = job.find_element_by_xpath('.//*[@class="salaryList"]').text.replace(' ', '').replace('\n', '')
            address = job.find_element_by_xpath('.//*[@class="cityConJobsWork"]').text.replace(' ', '').replace('\n', '')
            company_name = job.find_element_by_xpath('.//*[@class="jobRight"]//dt/a').text.replace(' ', '').replace('\n', '')
            if len(job.find_elements_by_xpath('.//*[@class="company-info"]/span')) == 3:
                company_type = job.find_element_by_xpath('.//*[@class="company-info"]/span[1]').text.replace(' ', '').replace('\n', '')
                company_size = job.find_element_by_xpath('.//*[@class="company-info"]/span[2]').text.replace(' ', '').replace('\n', '')
                company_business = job.find_element_by_xpath('.//*[@class="company-info"]/span[3]').text.replace(' ', '').replace('\n', '')
            else:
                company_type = '-'
                company_size = '-'
                company_business = '-'
            try:
                other = job.find_element_by_xpath('.//*[@class="hrBox_r_text"]').text
            except:
                other = '-'
            refresh_time = job.find_element_by_xpath('.//*[@class="time"]').text.replace(' ', '').replace('\n', '').replace('刷新过', '')
            worksheet.cell(row=rows, column=1, value=job_name)
            worksheet.cell(row=rows, column=2, value=salary)
            worksheet.cell(row=rows, column=3, value=address)
            worksheet.cell(row=rows, column=4, value=company_name)
            worksheet.cell(row=rows, column=5, value=company_type)
            worksheet.cell(row=rows, column=6, value=company_size)
            worksheet.cell(row=rows, column=7, value=company_business)
            worksheet.cell(row=rows, column=8, value=other)
            worksheet.cell(row=rows, column=9, value=refresh_time)
            data['职位名称'].append(job_name)
            data['职位月薪'].append(salary)
            data['工作地点'].append(address)
            data['公司名称'].append(company_name)
            data['公司类型'].append(company_type)
            data['公司规模'].append(company_size)
            data['经营项目'].append(company_business)
            data['其他信息'].append(other)
            data['刷新日期'].append(refresh_time)
            print(f'{job_name}   {salary}   {address}   {company_name}   {company_type}   {company_size}   {company_business}   {other}   {refresh_time}')
            rows += 1
        workbook.save('job.xlsx')
    pd_data = pd.DataFrame(data)
    create_line_chart(pd_data)
    create_pie_chart(pd_data)
    create_word_chart(pd_data)
    print('爬取完毕')
