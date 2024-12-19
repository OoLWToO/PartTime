import time

import requests
from lxml import etree
import openpyxl

# 设置请求头
headers = {
    'accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9',
    'accept-encoding': 'gzip, deflate, br',
    'accept-language': 'zh-CN,zh;q=0.9',
    'cookie': 'Hm_lvt_e9daa3d2d6076d2d8ec79d05f050c0d5=1716811706,1716825458; Hm_lvt_029c70397ba7bba8caeb29017c83b8d8=1716811706,1716825458; Hm_lpvt_029c70397ba7bba8caeb29017c83b8d8=1716825462; Hm_lpvt_e9daa3d2d6076d2d8ec79d05f050c0d5=1716825462',
    'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/107.0.0.0 Safari/537.36',
}

# 发送请求
url = 'https://bus.mapbar.com/xuzhou/xianlu/'
response = requests.get(url, headers=headers)
html = etree.HTML(response.text)

# 初始化Excel
workbook = openpyxl.Workbook()
worksheet = workbook.active
worksheet.column_dimensions['A'].width = 27
worksheet.column_dimensions['B'].width = 45
worksheet.column_dimensions['C'].width = 30
worksheet.column_dimensions['D'].width = 30
worksheet.cell(row=1, column=1, value='线路名称')
worksheet.cell(row=1, column=2, value='所属公司')
worksheet.cell(row=1, column=3, value='起点站首末车时间')
worksheet.cell(row=1, column=4, value='终点站站首末车时间')
rows = 2

# 获取所有路线详细的链接
bus_urls = html.xpath('//*[@class="ChinaTxt"]/dd/a/@href')
for url in bus_urls:
    # 请求路线详细链接
    b_response = requests.get(url, headers=headers)
    b_html = etree.HTML(b_response.text)
    # 根据xpath获取元素
    bus_name = b_html.xpath('//*[@class="publicBox"]//ul[@class="clr"]/li[1]/text()')
    bus_company = b_html.xpath('//*[@class="publicBox"]//ul[@class="clr"]/li[4]/text()')
    start_point_time = b_html.xpath('//*[@class="publicBox"]//ul[@class="clr"]/p[contains(text(),"起点站")]/text()')
    end_point_time = b_html.xpath('//*[@class="publicBox"]//ul[@class="clr"]/p[contains(text(),"终点站")]/text()')
    # 检查元素是否为空
    bus_name = check_empty(bus_name)
    bus_company = check_empty(bus_company)
    start_point_time = check_empty(start_point_time)
    end_point_time = check_empty(end_point_time)
    # 格式化数据
    start_point_time = start_point_time.replace('起点站首末车时间:', '')
    end_point_time = end_point_time.replace('终点站首末车时间:', '')
    # 存入Excel
    worksheet.cell(row=rows, column=1, value=bus_name)
    worksheet.cell(row=rows, column=2, value=bus_company)
    worksheet.cell(row=rows, column=3, value=start_point_time)
    worksheet.cell(row=rows, column=4, value=end_point_time)
    rows += 1
    print(f'{bus_name}   {bus_company}   {start_point_time}   {end_point_time}')
    # time.sleep一下不然容易被封ip，可以设置长一点
    time.sleep(.5)
# 保存Excel
workbook.save('徐州巴士信息.xlsx')
print('爬取完毕')