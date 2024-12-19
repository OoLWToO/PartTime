import re
import time

from openpyxl import Workbook

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC


def create_driver():
    chrome_driver_path = '../driver/chromedriver_119.exe'
    option = webdriver.ChromeOptions()
    option.add_experimental_option("detach", True)
    driver = webdriver.Chrome(executable_path=chrome_driver_path, chrome_options=option)
    driver.implicitly_wait(5)
    driver.maximize_window()
    return driver


if __name__ == "__main__":
    username = '17801010985@163.com'
    password = '123456'
    driver = create_driver()
    driver.get('https://pro.imdb.com/login/ap?u=/login/lwa&imdbPageAction=signUp')
    # 登录操作
    # IMDb登录
    # driver.find_element_by_xpath('//span[contains(text(),"IMDb")]').click()
    # driver.find_element_by_xpath('//*[@id="ap_email"]').send_keys(username)
    # driver.find_element_by_xpath('//*[@id="ap_password"]').send_keys(password)
    # driver.find_element_by_xpath('//*[@id="signInSubmit"]').click()
    # 亚马逊登录
    driver.find_element_by_xpath('(//span[contains(text(),"Amazon")])[2]').click()
    driver.find_element_by_xpath('//*[@id="ap_email"]').send_keys(username)
    driver.find_element_by_xpath('//*[@class="a-button-inner"]').click()
    driver.find_element_by_xpath('//*[@id="ap_password"]').send_keys(password)
    driver.find_element_by_xpath('//*[@id="signInSubmit"]').click()
    time.sleep(5)
    workbook = Workbook()
    for year in range(2019, 2020):
        driver.get(f'https://pro.imdb.com/boxoffice/year/world/{year}/')
        if year == 2019:
            worksheet = workbook.active
            worksheet.title = f'{year}'
        else:
            worksheet = workbook.create_sheet(title=f'{year}')
        worksheet.cell(row=1, column=1, value='电影名')
        worksheet.cell(row=1, column=2, value='公司')
        worksheet.cell(row=1, column=3, value='国家')
        rows = 2
        for i in range(200):
            while True:
                try:
                    driver.find_element_by_xpath(f'(//*[@id="table"]//tr[td[2]]/td[2]/a)[{i + 1}]').click()
                    break
                except:
                    driver.refresh()
            try:
                name = driver.find_element_by_xpath('//*[@class="a-size-extra-large"] | //*[@id="title_heading"]//*[@class="a-size-unset"]/a').text
            except:
                driver.refresh()
                try:
                    name = driver.find_element_by_xpath('//*[@class="a-size-extra-large"] | //*[@id="title_heading"]//*[@class="a-size-unset"]/a').text
                except:
                    driver.refresh()
                    name = driver.find_element_by_xpath('//*[@class="a-size-extra-large"] | //*[@id="title_heading"]//*[@class="a-size-unset"]/a').text
            try:
                company = driver.find_element_by_xpath('(//*[@id="contacts"]//*[@class="a-row a-spacing-mini a-spacing-top-none"])[1]//a').text
            except:
                company = ''
            try:
                country = driver.find_element_by_xpath('//span[contains(text(),"Production Company")]/../../following-sibling::div//li[position() = last()-1]').text
            except:
                country = ''
            print(f'{name}   {company}   {country}')
            driver.back()
            worksheet.cell(row=rows, column=1, value=name)
            worksheet.cell(row=rows, column=2, value=company)
            worksheet.cell(row=rows, column=3, value=country)
            rows += 1
        workbook.save('2009年-2019年top200影片公司及国家.xlsx')
