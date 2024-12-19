import time

import requests
from lxml import etree
from openpyxl.workbook import Workbook

headers = {
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7',
    'Accept-Encoding': 'gzip, deflate, br, zstd',
    'accept-language': 'zh-CN,zh;q=0.9',
    'Cache-Control': 'max-age=0',
    'Cookie': '__jsluid_s=c9bf2381199315a3f9d70fab5efbea9b; __jsl_clearance_s=1718590950.375|0|BNVMH65p0N9580f3nELVxfBlj1g%3D; PHPSESSID=mnnnrrlijttt606feek3bkdvg0; mfw_uuid=666f9de7-be92-5b8a-42db-f09d461618a6; oad_n=a%3A3%3A%7Bs%3A3%3A%22oid%22%3Bi%3A1029%3Bs%3A2%3A%22dm%22%3Bs%3A15%3A%22www.mafengwo.cn%22%3Bs%3A2%3A%22ft%22%3Bs%3A19%3A%222024-06-17+10%3A22%3A31%22%3B%7D; __mfwc=direct; __mfwa=1718590951232.55084.1.1718590951232.1718590951232; __mfwlv=1718590951; __mfwvn=1; Hm_lvt_8288b2ed37e5bc9b4c9f7008798d2de0=1718590951; uva=s%3A150%3A%22a%3A3%3A%7Bs%3A2%3A%22lt%22%3Bi%3A1718590952%3Bs%3A10%3A%22last_refer%22%3Bs%3A82%3A%22https%3A%2F%2Fwww.mafengwo.cn%2Flocaldeals%2F0-0-M10132-0-0-0-0-0.html%3Ffrom%3Dlocaldeals_index%22%3Bs%3A5%3A%22rhost%22%3BN%3B%7D%22%3B; __mfwurd=a%3A3%3A%7Bs%3A6%3A%22f_time%22%3Bi%3A1718590952%3Bs%3A9%3A%22f_rdomain%22%3Bs%3A15%3A%22www.mafengwo.cn%22%3Bs%3A6%3A%22f_host%22%3Bs%3A3%3A%22www%22%3B%7D; __mfwuuid=666f9de7-be92-5b8a-42db-f09d461618a6; __mfwb=7d6d46205828.2.direct; __mfwlt=1718590972; Hm_lpvt_8288b2ed37e5bc9b4c9f7008798d2de0=1718590973; bottom_ad_status=0',
    'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/119.0.0.0 Safari/537.36',
}

if __name__ == '__main__':
    workbook = Workbook()
    for year in range(2016, 2020):
        if year == 2016:
            worksheet = workbook.active
            worksheet.title = f'{year}'
        else:
            worksheet = workbook.create_sheet(title=f'{year}')
        rows = 2
        detail_url = []
        # 请求地址
        url = f'https://www.boxofficemojo.com/year/world/{year}/'
        r = requests.get(url, headers=headers)
        html = etree.HTML(r.text)
        movie_list = html.xpath('//*[@id="table"]//tr[td[2]]')
        for movie in movie_list:
            detail_url.append(movie.xpath('./td[2]/a/@href')[0])
            rows += 1
        worksheet.cell(row=1, column=1, value='电影名')
        worksheet.cell(row=1, column=2, value='国家')
        worksheet.cell(row=1, column=3, value='总票房')
        rows = 2
        for url in detail_url:
            print(f'https://www.boxofficemojo.com{url}')
            r = requests.get(f'https://www.boxofficemojo.com{url}', headers=headers, timeout=10)
            time.sleep(.5)
            html = etree.HTML(r.text)
            name = html.xpath('//*[@class="a-size-extra-large"]/text()')[0]
            box_office_list = html.xpath('//tr[td[2]]')
            for box_office in box_office_list:
                country = box_office.xpath('./td[1]/a/text()')[0]
                try:
                    gross = box_office.xpath('./td[4]/span/text()')[0]
                except:
                    gross = ''
                print(f'{year}   {name}   {country}   {gross}')
                worksheet.cell(row=rows, column=1, value=name)
                worksheet.cell(row=rows, column=2, value=country)
                worksheet.cell(row=rows, column=3, value=gross)
                rows += 1
            workbook.save('2009年-2019年top200影片各个国家票房.xlsx')