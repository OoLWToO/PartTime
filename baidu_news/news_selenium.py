from datetime import datetime

import pandas as pd
from openpyxl import Workbook

from selenium import webdriver

# 这是完整的关键词，如果程序运行过程中中断了，就把已经跑了的关键词删掉再跑，跑之前把已经保存的Excel放到其他地方，不然会被覆盖，跑完之后再整合到一起
# keywords1 = ['农业银行', '交通银行', '工商银行', '建设银行', '中国银行', '邮储银行', '平安银行', '浦发银行', '华夏银行',
#              '民生银行', '招商银行', '兴业银行', '光大银行', '浙商银行', '中信银行', '兰州银行', '西安银行', '南京银行',
#              '郑州银行',
#              '青岛银行',
#              '苏州银行', '江苏银行', '杭州银行', '北京银行', '厦门银行', '上海银行', '齐鲁银行', '长沙银行', '成都银行',
#              '重庆银行',
#              '贵阳银行', '宁波银行', '江阴银行', '江阴农商银行', '张家港行', '张家港农村商业银行', '青农商行',
#              '青岛农商银行',
#              '无锡银行', '无锡农村商业银行', '渝农商行', '重庆农村商业银行    ', '常熟银行', '常熟农商银行', '瑞丰银行',
#              '沪农商行', '上海农商银行', '紫金银行', '紫金农商银行']

keywords1 = ['沪农商行', '上海农商银行', '紫金银行', '紫金农商银行']

keywords2 = ['智能投顾', '智能营销', '智能客服', '智能柜台', '数字营销', '手机银行', 'APP', '移动支付', '手机支付',
             '第三方支付', '数字信贷', '预测模型', '大数据风控',
             '行为建模', '评分模型', '反欺诈模型', '多方安全计算', '联邦学习', '差分隐私技术', '联邦链', '风险知识图谱',
             '资金流动监测', '操作风险', '模型和算法风险',
             '金融科技创新', '专利申请', '专利授权', '软件著作权', '自研系统', '金融科技国际标准', '大数据', '云计算',
             '区块链', '人工智能', '信息安全', '物联网', '5G',
             'API', '开放银行', '混合现实', '增强现实']

data = {
    '关键词1': [],
    '关键词2': [],
    '标题': [],
    '时间': [],
    '摘要': []
}


def create_driver():
    chrome_driver_path = '../driver/chromedriver_119.exe'
    option = webdriver.ChromeOptions()
    option.add_experimental_option("detach", True)
    driver = webdriver.Chrome(executable_path=chrome_driver_path, chrome_options=option)
    driver.implicitly_wait(.5)
    driver.maximize_window()
    return driver


# 精确查询函数
def check_accuracy(word, title, abstract):
    if word in ['智能投顾', '智能营销', '智能客服', '智能柜台', '专利申请', '专利授权']:
        if word[2:] in title or word[2:] in abstract:
            return True
        else:
            return False
    if word in ['移动支付', '手机支付', '预测模型', '评分模型', '混合现实', '增强现实']:
        if word[:2] in title or word[:2] in abstract:
            return True
        else:
            return False
    if word in ['第三方支付', '反欺诈模型']:
        if word[:3] in title or word[:3] in abstract:
            return True
        else:
            return False
    return True


if __name__ == "__main__":
    # 初始化Excel
    workbook = Workbook()
    worksheet = workbook.active
    rows = 1
    columns = 3
    for word2 in keywords2:
        worksheet.cell(row=rows, column=columns, value=word2)
        columns += 1
    rows += 1
    columns = 1
    for word1 in keywords1:
        for year in range(2023, 2011, -1):
            worksheet.cell(row=rows, column=1, value=word1)
            worksheet.cell(row=rows, column=2, value=year)
            rows += 1
    # 创建driver
    driver = create_driver()
    driver.get('https://www.baidu.com/s?tn=news&rtt=4&bsst=1&cl=2&wd=%E6%B5%8B%E8%AF%95&medium=0')
    for word1 in keywords1:
        columns = 3
        for word2 in keywords2:
            rows = keywords1.index(word1) * 12 + 2
            # 退出标记，用于判断时间是否超出规定范围
            break_sign = False
            news_num = [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]
            # 搜索操作
            try:
                driver.find_element_by_xpath('//p[text()="百度安全验证"]')
                input('有验证码，手动操作一下！！！')
            except:
                pass
            driver.find_element_by_xpath('//*[@id="kw"]').clear()
            driver.find_element_by_xpath('//*[@id="kw"]').send_keys(f'{word1}{word2}')
            driver.find_element_by_xpath('//*[@id="su"]').click()
            # 获取咨询
            while True:
                try:
                    driver.find_element_by_xpath('//p[text()="百度安全验证"]')
                    input('有验证码，手动操作一下！！！')
                except:
                    pass
                news_list = driver.find_elements_by_xpath('//*[@id="content_left"]/div[div]')
                for news in news_list:
                    try:
                        time = news.find_element_by_xpath('.//span[contains(@class,"c-color-gray2")]').text
                    except:
                        time = ''
                    if '年' in time:
                        if datetime(2012, 1, 1) <= datetime.strptime(time, '%Y年%m月%d日') <= datetime(2023, 12, 31):
                            title = news.find_element_by_xpath('.//a[contains(@class,"news-title")]').text
                            abstract = news.find_element_by_xpath('.//span[contains(@aria-label,"摘要")]').text
                            if check_accuracy(word2, title, abstract):
                                print(f'{word1}   {word2}   {title}   {time}   {abstract}')
                                time_year = datetime.strptime(time, '%Y年%m月%d日').year
                                # 统计咨询条数
                                if time_year == 2023:
                                    news_num[0] += 1
                                if time_year == 2022:
                                    news_num[1] += 1
                                if time_year == 2021:
                                    news_num[2] += 1
                                if time_year == 2020:
                                    news_num[3] += 1
                                if time_year == 2019:
                                    news_num[4] += 1
                                if time_year == 2018:
                                    news_num[5] += 1
                                if time_year == 2017:
                                    news_num[6] += 1
                                if time_year == 2016:
                                    news_num[7] += 1
                                if time_year == 2015:
                                    news_num[8] += 1
                                if time_year == 2014:
                                    news_num[9] += 1
                                if time_year == 2013:
                                    news_num[10] += 1
                                if time_year == 2012:
                                    news_num[11] += 1
                                data['关键词1'].append(word1)
                                data['关键词2'].append(word2)
                                data['标题'].append(title)
                                data['时间'].append(time)
                                data['摘要'].append(abstract)
                        # 判断年份是否超出范围
                        if datetime.strptime(time, '%Y年%m月%d日') < datetime(2012, 1, 1):
                            break_sign = True
                            break
                # 保存数据并退出循环进行下一个关键词查询
                if break_sign:
                    df = pd.DataFrame(data)
                    df.to_excel('百度咨询数据.xlsx', index=False)
                    for num in news_num:
                        worksheet.cell(row=rows, column=columns, value=num)
                        rows += 1
                    columns += 1
                    workbook.save('百度咨询数据统计.xlsx')
                    break
                else:
                    try:
                        driver.find_element_by_xpath('//a[@class="n" and contains(text(),"下一页")]').click()
                    except:
                        df = pd.DataFrame(data)
                        df.to_excel('百度咨询数据.xlsx', index=False)
                        for num in news_num:
                            worksheet.cell(row=rows, column=columns, value=num)
                            rows += 1
                        columns += 1
                        workbook.save('百度咨询数据统计.xlsx')
                        break
