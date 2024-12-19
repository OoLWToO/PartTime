import requests
from lxml import etree
from openpyxl import Workbook
import matplotlib.pyplot as plt


headers = {
    'accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9',
    'accept-encoding': 'gzip, deflate, br',
    'accept-language': 'zh-CN,zh;q=0.9',
    'cookie': 'Hm_lvt_e9daa3d2d6076d2d8ec79d05f050c0d5=1716811706,1716825458; Hm_lvt_029c70397ba7bba8caeb29017c83b8d8=1716811706,1716825458; Hm_lpvt_029c70397ba7bba8caeb29017c83b8d8=1716825462; Hm_lpvt_e9daa3d2d6076d2d8ec79d05f050c0d5=1716825462',
    'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/107.0.0.0 Safari/537.36',
}

workbook = Workbook()
worksheet = workbook.active
# 设置单元格宽度
worksheet.column_dimensions['A'].width = 30
worksheet.column_dimensions['B'].width = 14
worksheet.column_dimensions['C'].width = 14
worksheet.column_dimensions['D'].width = 18
worksheet.column_dimensions['E'].width = 40
worksheet.column_dimensions['F'].width = 12
worksheet.column_dimensions['G'].width = 12
worksheet.column_dimensions['H'].width = 30
worksheet.column_dimensions['I'].width = 10
worksheet.cell(row=1, column=1, value='职位名称')
worksheet.cell(row=1, column=2, value='职位月薪')
worksheet.cell(row=1, column=3, value='职位月薪')
worksheet.cell(row=1, column=4, value='工作地区')
worksheet.cell(row=1, column=5, value='公司名称')
worksheet.cell(row=1, column=6, value='公司类型')
worksheet.cell(row=1, column=7, value='公司规模')
worksheet.cell(row=1, column=8, value='经营项目')
worksheet.cell(row=1, column=9, value='刷新日期')
rows = 2
# 公司类型饼状图数据
company_type_data = []
company_type_value = []
# 公司规模饼状图数据
company_size_data = []
company_size_value = []
for page in range(1, 11):
    url = f'https://www.job001.cn/jobs?keyType=0&keyWord=&jobTypeId=&jobType=%E8%81%8C%E4%BD%8D%E7%B1%BB%E5%9E%8B&industry=&industryname=%E8%A1%8C%E4%B8%9A%E7%B1%BB%E5%9E%8B&workId=25.296&workPlace=%E4%BD%9B%E5%B1%B1%E5%B8%82&salary=&salaryType=&entType=&experience=&education=&entSize=&benefits=&reftime=&workTypeId=&sortField=&pageNo={page}&curItem=&searchType=1'
    r = requests.get(url, headers=headers)
    html = etree.HTML(r.text)
    jobs_list = html.xpath('//*[@class="jobsList"]')
    for job in jobs_list:
        job_name = job.xpath('.//*[@class="jobNameCon"]')[0].text.replace(' ', '').replace('\n', '')
        salary = job.xpath('.//*[@class="salaryList"]')[0].text.replace(' ', '').replace('\n', '')
        educational = job.xpath('.//*[@class="jobLeft"]/dl/dd[1]/text()')[2].replace(' ', '').replace('\n', '')
        address = job.xpath('.//*[@class="cityConJobsWork"]')[0].text.replace(' ', '').replace('\n', '')
        company_name = job.xpath('.//*[@class="jobRight"]//dt/a')[0].text.replace(' ', '').replace('\n', '')
        company_type = job.xpath('.//*[@class="company-info"]/span[1]')[0].text.replace(' ', '').replace('\n', '')
        company_size = job.xpath('.//*[@class="company-info"]/span[2]')[0].text.replace(' ', '').replace('\n', '')
        try:
            company_business = job.xpath('.//*[@class="company-info"]/span[3]')[0].text.replace(' ', '').replace('\n', '')
        except:
            company_business = '-'
        refresh_time = job.xpath('.//*[@class="time"]')[0].text.replace(' ', '').replace('\n', '').replace('刷新过', '')
        print(f'{job_name}   {salary}   {educational}   {address}   {company_name}   {company_type}   {company_size}   {company_business}   {refresh_time}')
        worksheet.cell(row=rows, column=1, value=job_name)
        worksheet.cell(row=rows, column=2, value=salary)
        worksheet.cell(row=rows, column=3, value=educational)
        worksheet.cell(row=rows, column=4, value=address)
        worksheet.cell(row=rows, column=5, value=company_name)
        worksheet.cell(row=rows, column=6, value=company_type)
        worksheet.cell(row=rows, column=7, value=company_size)
        worksheet.cell(row=rows, column=8, value=company_business)
        worksheet.cell(row=rows, column=9, value=refresh_time)
        rows += 1
        # 统计数据
        if company_type not in company_type_data:
            company_type_data.append(company_type)
            company_type_value.append(1)
        else:
            company_type_value[company_type_data.index(company_type)] += 1
        if company_size not in company_size_data:
            company_size_data.append(company_size)
            company_size_value.append(1)
        else:
            company_size_value[company_size_data.index(company_size)] += 1
workbook.save('rencai.xlsx')

# 数据可视化
plt.rcParams['font.sans-serif'] = ['SimHei']
plt.rcParams['axes.unicode_minus'] = False
# 公司类型
plt.pie(company_type_value, labels=company_type_data, autopct='%1.1f%%', startangle=140)
plt.title('公司类型')
plt.axis('equal')
plt.show()
# 公司规模
plt.pie(company_size_value, labels=company_size_data, autopct='%1.1f%%', startangle=140)
plt.title('公司规模')
plt.axis('equal')
plt.show()







