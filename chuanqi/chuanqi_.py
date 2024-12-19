import openpyxl
import requests


# 全局配置变量
strorigin = 'CNSHA'
strdest = 'nlrtm',
file_name = '上海 鹿特丹'
workbook = openpyxl.Workbook()
worksheet = workbook.active
rows = 2


def get_data(strorigin, strdest, page):
    url = 'https://ezocean.com/SCHEDULE/FnRetrieveScheduleList'
    headers = {
        'accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9',
        'accept-encoding': 'gzip, deflate, br',
        'accept-language': 'zh-CN,zh;q=0.9',
        'cookie': 'Hm_lvt_e9daa3d2d6076d2d8ec79d05f050c0d5=1716811706,1716825458; Hm_lvt_029c70397ba7bba8caeb29017c83b8d8=1716811706,1716825458; Hm_lpvt_029c70397ba7bba8caeb29017c83b8d8=1716825462; Hm_lpvt_e9daa3d2d6076d2d8ec79d05f050c0d5=1716825462',
        'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/107.0.0.0 Safari/537.36',
    }
    r = requests.post(url, data=data, headers=headers)
    js_lst = r.json()
    return js_lst['schedulers']


def write_in_excel(data_list):
    global rows
    for ship in data_list:
        worksheet.cell(row=rows, column=1, value=ship['ServiceRefenceID'])
        worksheet.cell(row=rows, column=2, value=ship['Carriers'])
        worksheet.cell(row=rows, column=3, value=ship['VesselName'])
        worksheet.cell(row=rows, column=4, value=ship['Voyage'])
        worksheet.cell(row=rows, column=5, value=ship['OriginName'])
        worksheet.cell(row=rows, column=6, value=ship['DestinationName'])
        worksheet.cell(row=rows, column=7, value=ship['OriginTerminal'])
        worksheet.cell(row=rows, column=8, value=ship['DestinationTerminal'])
        worksheet.cell(row=rows, column=9, value=ship['FromETD'])
        worksheet.cell(row=rows, column=10, value=ship['fromSchedule'])
        worksheet.cell(row=rows, column=11, value=ship['ToETA'])
        worksheet.cell(row=rows, column=12, value=ship['toSchedule'])
        worksheet.cell(row=rows, column=13, value=ship['consortium'])
        worksheet.cell(row=rows, column=14, value=ship['Operator'])
        worksheet.cell(row=rows, column=15, value=ship['VesselCapacity'])
        worksheet.cell(row=rows, column=16, value=ship['TransitTime'])
        rows += 1


if __name__ == '__main__':
    worksheet.column_dimensions['A'].width = 5
    worksheet.column_dimensions['B'].width = 52
    worksheet.column_dimensions['C'].width = 32
    worksheet.column_dimensions['D'].width = 10
    worksheet.column_dimensions['E'].width = 15
    worksheet.column_dimensions['F'].width = 15
    worksheet.column_dimensions['G'].width = 10
    worksheet.column_dimensions['H'].width = 10
    worksheet.column_dimensions['I'].width = 12
    worksheet.column_dimensions['K'].width = 12
    worksheet.column_dimensions['N'].width = 12
    worksheet.cell(row=1, column=1, value='ID')
    worksheet.cell(row=1, column=2, value='船公司（航线代码）')
    worksheet.cell(row=1, column=3, value='船名')
    worksheet.cell(row=1, column=4, value='航次')
    worksheet.cell(row=1, column=5, value='起运港')
    worksheet.cell(row=1, column=6, value='目的港')
    worksheet.cell(row=1, column=7, value='起运码头')
    worksheet.cell(row=1, column=8, value='目的码头')
    worksheet.cell(row=1, column=9, value='起运日期')
    worksheet.cell(row=1, column=10, value='起运延误')
    worksheet.cell(row=1, column=11, value='到港日期')
    worksheet.cell(row=1, column=12, value='到港延误')
    worksheet.cell(row=1, column=13, value='联盟')
    worksheet.cell(row=1, column=14, value='船舶经营人')
    worksheet.cell(row=1, column=15, value='装载量')
    worksheet.cell(row=1, column=16, value='航程')
    for page in range(1, 100):
        data_list = get_data(strorigin, strdest, page)
        if not data_list:
            break
        write_in_excel(data_list)
    workbook.save(f'{file_name}.xlsx')
    print('爬取完毕')