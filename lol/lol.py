import re
from openpyxl import Workbook

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

rows = 0


# 构建chrome_driver
def create_driver():
    # chromedriver最新版下载地址
    # https://googlechromelabs.github.io/chrome-for-testing/#stable
    chrome_driver_path = "./chromedriver_119.exe"
    options = Options()
    options.add_argument("--start-maximized")
    c_driver = webdriver.Chrome(executable_path=chrome_driver_path, options=options)
    c_driver.implicitly_wait(5)
    c_driver.get('https://www.esports8.com/lol')
    c_driver.maximize_window()
    return c_driver


# 获取图片链接
def get_image_url(driver, xpath):
    player_logo = re.search(r"url\(\"([^)]+)\"\)", driver.find_element(By.XPATH, xpath).get_attribute('style'))
    if player_logo:
        player_logo = player_logo.group(1)
        return player_logo
    return '无图片'


# 写入表格，默认增行1
def sheet_row_add(worksheet, row, column, value, add=1):
    global rows
    rows = row
    worksheet.cell(row=row, column=column, value=value)
    rows += add


if __name__ == "__main__":
    driver = create_driver()
    workbook = Workbook()
    match_eles = driver.find_elements(By.XPATH, '//*[@class="match-list-item"]')
    # 前瞻分析url
    ahead_url = [f'{match.get_attribute("href")}/analysis' for match in match_eles]
    # 阵容体系url
    lineup_url = [f'{match.get_attribute("href")}/players' for match in match_eles]
    # 前瞻分析
    for i in range(len(ahead_url)):
        driver.get(ahead_url[i])
        # 历史交锋
        team_A_logo = get_image_url(driver, '//*[@class="match-info"]//div[@class="team-logo-wrap left"]/div')
        team_A_name = driver.find_element(By.XPATH, '//*[@class="match-info"]//div[@class="team-logo-wrap left"]/div[@class="team-name"]').text
        team_B_logo = get_image_url(driver, '//*[@class="match-info"]//div[@class="team-logo-wrap right"]/div')
        team_B_name = driver.find_element(By.XPATH, '//*[@class="match-info"]//div[@class="team-logo-wrap right"]/div[@class="team-name"]').text

        # 初始化excel
        rows = 1
        if i == 0:
            worksheet = workbook.active  # 第一个子表
            worksheet.title = f'{team_A_name} VS {team_B_name}'
        else:
            worksheet = workbook.create_sheet(title=f'{team_A_name} VS {team_B_name}')  # 其他子表

        sheet_row_add(worksheet=worksheet, row=rows, column=1, value='比赛链接')
        sheet_row_add(worksheet=worksheet, row=rows, column=1, value=ahead_url[i], add=2)
        sheet_row_add(worksheet=worksheet, row=rows, column=1, value='比赛队伍')
        sheet_row_add(worksheet=worksheet, row=rows, column=1, value=team_A_logo, add=0)
        sheet_row_add(worksheet=worksheet, row=rows, column=2, value=team_B_logo)
        sheet_row_add(worksheet=worksheet, row=rows, column=1, value=team_A_name, add=0)
        sheet_row_add(worksheet=worksheet, row=rows, column=2, value=team_B_name, add=2)
        sheet_row_add(worksheet=worksheet, row=rows, column=1, value='历史交锋')
        history_box = driver.find_element(By.XPATH, '//*[contains(@class,"historyBox")]')
        if '暂无数据' in history_box.text:
            sheet_row_add(worksheet=worksheet, row=rows, column=1, value='暂无历史交锋')
            rows += 1
        else:
            history_box.find_element(By.XPATH, './/span[text()="近"]/following-sibling::div').click()
            list_btn = WebDriverWait(driver, 3).until(EC.presence_of_all_elements_located((By.XPATH, '//*[@class="el-scrollbar"]//li/span[text()="50"]')))
            for btn in list_btn:
                if btn.is_displayed():
                    btn.click()
            # 历史对局
            sheet_row_add(worksheet=worksheet, row=rows, column=1, value='历史对局')
            sheet_row_add(worksheet=worksheet, row=rows, column=1, value='赛事', add=0)
            sheet_row_add(worksheet=worksheet, row=rows, column=2, value='时间', add=0)
            sheet_row_add(worksheet=worksheet, row=rows, column=3, value='击杀', add=0)
            sheet_row_add(worksheet=worksheet, row=rows, column=4, value='时长', add=0)
            sheet_row_add(worksheet=worksheet, row=rows, column=5, value='一血', add=0)
            sheet_row_add(worksheet=worksheet, row=rows, column=6, value='一塔', add=0)
            sheet_row_add(worksheet=worksheet, row=rows, column=7, value='五杀', add=0)
            sheet_row_add(worksheet=worksheet, row=rows, column=8, value='十杀', add=0)
            sheet_row_add(worksheet=worksheet, row=rows, column=9, value='胜负')
            history_table = history_box.find_elements(By.XPATH, './/tbody/tr')
            for history in history_table:
                competition = history.find_element(By.XPATH, './td[1]').text
                time = history.find_element(By.XPATH, './td[2]').text
                kill = history.find_element(By.XPATH, './td[3]').text
                # 去除大比分获胜标记
                if '大' in kill:
                    kill = kill.replace('大\n', '')
                time_length = history.find_element(By.XPATH, './td[4]').text
                first_blood = history.find_element(By.XPATH, './td[5]//span[1]').get_attribute('class').endswith('g')
                first_tower = history.find_element(By.XPATH, './td[5]//span[2]').get_attribute('class').endswith('g')
                five_kills = history.find_element(By.XPATH, './td[6]//span[1]').get_attribute('class').endswith('g')
                ten_kills = history.find_element(By.XPATH, './td[6]//span[1]').get_attribute('class').endswith('g')
                win = history.find_element(By.XPATH, './td[7]//span').get_attribute('class').endswith('win')
                sheet_row_add(worksheet=worksheet, row=rows, column=1, value=competition, add=0)
                sheet_row_add(worksheet=worksheet, row=rows, column=2, value=time, add=0)
                sheet_row_add(worksheet=worksheet, row=rows, column=3, value=kill, add=0)
                sheet_row_add(worksheet=worksheet, row=rows, column=4, value=time_length, add=0)
                sheet_row_add(worksheet=worksheet, row=rows, column=5, value='是' if first_blood else '否', add=0)
                sheet_row_add(worksheet=worksheet, row=rows, column=6, value='是' if first_tower else '否', add=0)
                sheet_row_add(worksheet=worksheet, row=rows, column=7, value='是' if five_kills else '否', add=0)
                sheet_row_add(worksheet=worksheet, row=rows, column=8, value='是' if ten_kills else '否', add=0)
                sheet_row_add(worksheet=worksheet, row=rows, column=9, value='胜' if win else '负')
            rows += 1

            # 历史数据
            sheet_row_add(worksheet=worksheet, row=rows, column=1, value='历史数据')
            sheet_row_add(worksheet=worksheet, row=rows, column=1, value='队伍名', add=0)
            sheet_row_add(worksheet=worksheet, row=rows, column=2, value='胜率', add=0)
            sheet_row_add(worksheet=worksheet, row=rows, column=3, value='一血率', add=0)
            sheet_row_add(worksheet=worksheet, row=rows, column=4, value='一塔率', add=0)
            sheet_row_add(worksheet=worksheet, row=rows, column=5, value='五杀率', add=0)
            sheet_row_add(worksheet=worksheet, row=rows, column=6, value='十杀率', add=0)
            sheet_row_add(worksheet=worksheet, row=rows, column=7, value='时长>38', add=0)
            sheet_row_add(worksheet=worksheet, row=rows, column=8, value='击杀>40.5')
            history_wrap = history_box.find_element(By.XPATH, './/div[@class="history-battle-wrap"]')
            team_A_winning_rate = history_wrap.find_element(By.XPATH, './/*[@class="winRate"]/div[1]//div[@class="el-progress__text"]').text
            team_A_one_blood_rate = history_wrap.find_element(By.XPATH, './/*[@class="side-progress"][1]//div[@class="home-value"]').text
            team_A_one_tower_rate = history_wrap.find_element(By.XPATH, './/*[@class="side-progress"][2]//div[@class="home-value"]').text
            team_A_five_kill_rate = history_wrap.find_element(By.XPATH, './/*[@class="side-progress"][3]//div[@class="home-value"]').text
            team_A_ten_kill_rate = history_wrap.find_element(By.XPATH, './/*[@class="side-progress"][4]//div[@class="home-value"]').text
            team_B_winning_rate = history_wrap.find_element(By.XPATH, './/*[@class="winRate"]/div[3]//div[@class="el-progress__text"]').text
            team_B_one_blood_rate = history_wrap.find_element(By.XPATH, './/*[@class="side-progress"][1]//div[@class="away-value"]').text
            team_B_one_tower_rate = history_wrap.find_element(By.XPATH, './/*[@class="side-progress"][2]//div[@class="away-value"]').text
            team_B_five_kill_rate = history_wrap.find_element(By.XPATH, './/*[@class="side-progress"][3]//div[@class="away-value"]').text
            team_B_ten_kill_rate = history_wrap.find_element(By.XPATH, './/*[@class="side-progress"][4]//div[@class="away-value"]').text
            time_wrap = history_wrap.find_element(By.XPATH, './/*[@class="single-progress"][1]/p[2]').text
            kill_wrap = history_wrap.find_element(By.XPATH, './/*[@class="single-progress"][2]/p[2]').text
            sheet_row_add(worksheet=worksheet, row=rows, column=1, value=team_A_name, add=0)
            sheet_row_add(worksheet=worksheet, row=rows, column=2, value=team_A_winning_rate, add=0)
            sheet_row_add(worksheet=worksheet, row=rows, column=3, value=team_A_one_blood_rate, add=0)
            sheet_row_add(worksheet=worksheet, row=rows, column=4, value=team_A_one_tower_rate, add=0)
            sheet_row_add(worksheet=worksheet, row=rows, column=5, value=team_A_five_kill_rate, add=0)
            sheet_row_add(worksheet=worksheet, row=rows, column=6, value=team_A_ten_kill_rate, add=0)
            sheet_row_add(worksheet=worksheet, row=rows, column=7, value=time_wrap, add=0)
            sheet_row_add(worksheet=worksheet, row=rows, column=8, value=kill_wrap)
            sheet_row_add(worksheet=worksheet, row=rows, column=1, value=team_B_name, add=0)
            sheet_row_add(worksheet=worksheet, row=rows, column=2, value=team_B_winning_rate, add=0)
            sheet_row_add(worksheet=worksheet, row=rows, column=3, value=team_B_one_blood_rate, add=0)
            sheet_row_add(worksheet=worksheet, row=rows, column=4, value=team_B_one_tower_rate, add=0)
            sheet_row_add(worksheet=worksheet, row=rows, column=5, value=team_B_five_kill_rate, add=0)
            sheet_row_add(worksheet=worksheet, row=rows, column=6, value=team_B_ten_kill_rate, add=0)
            sheet_row_add(worksheet=worksheet, row=rows, column=7, value=time_wrap, add=0)
            sheet_row_add(worksheet=worksheet, row=rows, column=8, value=kill_wrap)
            rows += 1

        # 事件触发胜率
        sheet_row_add(worksheet=worksheet, row=rows, column=1, value='事件触发胜率')
        sheet_row_add(worksheet=worksheet, row=rows, column=1, value='队伍名称', add=0)
        sheet_row_add(worksheet=worksheet, row=rows, column=2, value='获得一血后胜率', add=0)
        sheet_row_add(worksheet=worksheet, row=rows, column=3, value='摧毁一塔后胜率', add=0)
        sheet_row_add(worksheet=worksheet, row=rows, column=4, value='击杀首小龙后胜率', add=0)
        sheet_row_add(worksheet=worksheet, row=rows, column=5, value='击杀首大龙后胜率', add=0)
        sheet_row_add(worksheet=worksheet, row=rows, column=6, value='击杀首峡谷先锋后胜率')
        team_A_triggered_1 = driver.find_element(By.XPATH, '//*[contains(@class,"event-box")]//div[@class="side-progress"][1]//div[@class="home-value"]').text
        team_A_triggered_2 = driver.find_element(By.XPATH, '//*[contains(@class,"event-box")]//div[@class="side-progress"][2]//div[@class="home-value"]').text
        team_A_triggered_3 = driver.find_element(By.XPATH, '//*[contains(@class,"event-box")]//div[@class="side-progress"][3]//div[@class="home-value"]').text
        team_A_triggered_4 = driver.find_element(By.XPATH, '//*[contains(@class,"event-box")]//div[@class="side-progress"][4]//div[@class="home-value"]').text
        team_A_triggered_5 = driver.find_element(By.XPATH, '//*[contains(@class,"event-box")]//div[@class="side-progress"][5]//div[@class="home-value"]').text
        team_B_triggered_1 = driver.find_element(By.XPATH, '//*[contains(@class,"event-box")]//div[@class="side-progress"][1]//div[@class="away-value"]').text
        team_B_triggered_2 = driver.find_element(By.XPATH, '//*[contains(@class,"event-box")]//div[@class="side-progress"][2]//div[@class="away-value"]').text
        team_B_triggered_3 = driver.find_element(By.XPATH, '//*[contains(@class,"event-box")]//div[@class="side-progress"][3]//div[@class="away-value"]').text
        team_B_triggered_4 = driver.find_element(By.XPATH, '//*[contains(@class,"event-box")]//div[@class="side-progress"][4]//div[@class="away-value"]').text
        team_B_triggered_5 = driver.find_element(By.XPATH, '//*[contains(@class,"event-box")]//div[@class="side-progress"][5]//div[@class="away-value"]').text
        sheet_row_add(worksheet=worksheet, row=rows, column=1, value=team_A_name, add=0)
        sheet_row_add(worksheet=worksheet, row=rows, column=2, value=team_A_triggered_1, add=0)
        sheet_row_add(worksheet=worksheet, row=rows, column=3, value=team_A_triggered_2, add=0)
        sheet_row_add(worksheet=worksheet, row=rows, column=4, value=team_A_triggered_3, add=0)
        sheet_row_add(worksheet=worksheet, row=rows, column=5, value=team_A_triggered_4, add=0)
        sheet_row_add(worksheet=worksheet, row=rows, column=6, value=team_A_triggered_5)
        sheet_row_add(worksheet=worksheet, row=rows, column=1, value=team_B_name, add=0)
        sheet_row_add(worksheet=worksheet, row=rows, column=2, value=team_B_triggered_1, add=0)
        sheet_row_add(worksheet=worksheet, row=rows, column=3, value=team_B_triggered_2, add=0)
        sheet_row_add(worksheet=worksheet, row=rows, column=4, value=team_B_triggered_3, add=0)
        sheet_row_add(worksheet=worksheet, row=rows, column=5, value=team_B_triggered_4, add=0)
        sheet_row_add(worksheet=worksheet, row=rows, column=6, value=team_B_triggered_5)
        rows += 1

        # 阵容体系
        driver.get(lineup_url[i])
        sheet_row_add(worksheet=worksheet, row=rows, column=1, value='阵容体系')
        job_btn = driver.find_elements(By.XPATH, '//*[@class="el-button-group"]/button')
        for job in job_btn:
            job.click()
            position = job.text
            team_xpath = [1, 3]
            for team in team_xpath:
                sheet_row_add(worksheet=worksheet, row=rows, column=1, value='队伍', add=0)
                sheet_row_add(worksheet=worksheet, row=rows, column=2, value=team_A_name if team == 1 else team_B_name, add=0)
                sheet_row_add(worksheet=worksheet, row=rows, column=4, value='位置', add=0)
                sheet_row_add(worksheet=worksheet, row=rows, column=5, value=position)
                sheet_row_add(worksheet=worksheet, row=rows, column=1, value='选手照片', add=0)
                sheet_row_add(worksheet=worksheet, row=rows, column=2, value='选手名字', add=0)
                sheet_row_add(worksheet=worksheet, row=rows, column=3, value='胜率', add=0)
                sheet_row_add(worksheet=worksheet, row=rows, column=4, value='胜场', add=0)
                sheet_row_add(worksheet=worksheet, row=rows, column=5, value='负场', add=0)
                sheet_row_add(worksheet=worksheet, row=rows, column=6, value='英雄1', add=0)
                sheet_row_add(worksheet=worksheet, row=rows, column=7, value='KDA', add=0)
                sheet_row_add(worksheet=worksheet, row=rows, column=8, value='英雄2', add=0)
                sheet_row_add(worksheet=worksheet, row=rows, column=9, value='KDA', add=0)
                sheet_row_add(worksheet=worksheet, row=rows, column=10, value='英雄3', add=0)
                sheet_row_add(worksheet=worksheet, row=rows, column=11, value='KDA')
                player_list = driver.find_elements(By.XPATH, f'//*[@class="playerRoleTitle"]/div[{team}]/span')
                for player in player_list[:-1] if team == 1 else player_list[1:]:
                    player.click()
                    player_logo = get_image_url(driver, f'//*[@class="cont"]/div[{team}]//div[contains(@class,"logo")]')
                    player_name = driver.find_element(By.XPATH, f'//*[@class="cont"]/div[{team}]//div[@class="playerName"]').text
                    win_grate = driver.find_element(By.XPATH, f'//*[@class="raceRecord"]/div[{team}]//div[@class="el-progress__text"]').text
                    victories = driver.find_element(By.XPATH, f'//*[@class="raceRecord"]/div[{team}]/*[contains(@class,"js-flex")]/span[1]').text
                    failures = driver.find_element(By.XPATH, f'//*[@class="raceRecord"]/div[{team}]/*[contains(@class,"js-flex")]/span[2]').text
                    hero_1_image = "无图片"
                    hero_1_KDA = "无英雄"
                    hero_2_image = "无图片"
                    hero_2_KDA = "无英雄"
                    hero_3_image = "无图片"
                    hero_3_KDA = "无英雄"
                    # 这里可能没有xpath，捕获跳过就行了
                    try:
                        hero_1_image = get_image_url(driver, f'//*[@class="raceRecord"]/div[{team}]/*[@class="heroWrap"]/div[1]/*[@class="heroIcon"]')
                        if '无图片' not in hero_1_image:
                            hero_1_KDA = driver.find_element(By.XPATH, f'//*[@class="raceRecord"]/div[{team}]/*[@class="heroWrap"]/div[1]/*[@class="heroDetail"]').text
                        hero_2_image = get_image_url(driver, f'//*[@class="raceRecord"]/div[{team}]/*[@class="heroWrap"]/div[2]/*[@class="heroIcon"]')
                        if '无图片' not in hero_2_image:
                            hero_2_KDA = driver.find_element(By.XPATH, f'//*[@class="raceRecord"]/div[{team}]/*[@class="heroWrap"]/div[2]/*[@class="heroDetail"]').text
                        hero_3_image = get_image_url(driver, f'//*[@class="raceRecord"]/div[{team}]/*[@class="heroWrap"]/div[3]/*[@class="heroIcon"]')
                        if '无图片' not in hero_3_image:
                            hero_3_KDA = driver.find_element(By.XPATH, f'//*[@class="raceRecord"]/div[{team}]/*[@class="heroWrap"]/div[3]/*[@class="heroDetail"]').text
                    except:
                        pass
                    sheet_row_add(worksheet=worksheet, row=rows, column=1, value=player_logo, add=0)
                    sheet_row_add(worksheet=worksheet, row=rows, column=2, value=player_name, add=0)
                    sheet_row_add(worksheet=worksheet, row=rows, column=3, value=win_grate, add=0)
                    sheet_row_add(worksheet=worksheet, row=rows, column=4, value=victories, add=0)
                    sheet_row_add(worksheet=worksheet, row=rows, column=5, value=failures, add=0)
                    sheet_row_add(worksheet=worksheet, row=rows, column=6, value=hero_1_image, add=0)
                    sheet_row_add(worksheet=worksheet, row=rows, column=7, value=hero_1_KDA, add=0)
                    sheet_row_add(worksheet=worksheet, row=rows, column=8, value=hero_2_image, add=0)
                    sheet_row_add(worksheet=worksheet, row=rows, column=9, value=hero_2_KDA, add=0)
                    sheet_row_add(worksheet=worksheet, row=rows, column=10, value=hero_3_image, add=0)
                    sheet_row_add(worksheet=worksheet, row=rows, column=11, value=hero_3_KDA, add=1)
            rows += 1

    workbook.save('lol.xlsx')
